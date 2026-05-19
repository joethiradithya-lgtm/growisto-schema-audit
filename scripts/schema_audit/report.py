"""
report.py — render results.json into an Excel workbook.

Output: outputs/schema-audit-<date>.xlsx with columns:
    A — Page URL
    B — Schemas Present       (one per line, in-cell newlines)
    C — Schemas Missing/Incomplete  (one per line, priority-tagged)
    D — JSON-LD Template      (paste-ready, full body inline; multiple schemas
                               separated by '———' divider, aligned to column C)

Used by run.py after the audit pipeline. Can also be invoked standalone:
    python report.py results.json
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


HERE = Path(__file__).parent
OUT_DIR = HERE / "outputs"
DIVIDER = "\n\n———\n\n"


def _row_for(result: dict) -> tuple[str, str, str, str]:
    """Build the (A, B, C, D) tuple for one URL result."""
    url = result.get("url", result.get("input_url", ""))

    if result.get("error"):
        return (url, f"ERROR: {result['error']}", "", "")

    present = "\n".join(result.get("present", [])) or "(none)"

    missing = result.get("missing", [])
    issues = result.get("issues", [])
    incomplete = [i for i in issues if i.get("status") == "INCOMPLETE"]

    missing_lines: list[str] = [f"{m['type']} [{m['priority']}]" for m in missing]
    for inc in incomplete:
        props = inc.get("missing_properties", [])
        suffix = f" — missing: {', '.join(props)}" if props else ""
        missing_lines.append(f"{inc['type']} [INCOMPLETE]{suffix}")
    missing_str = "\n".join(missing_lines) or "(none)"

    templates_map = result.get("templates", {})
    template_blocks: list[str] = []
    for m in missing:
        tpl = templates_map.get(m["type"], f"// no template available for {m['type']}")
        template_blocks.append(tpl)
    for inc in incomplete:
        tpl = templates_map.get(
            inc["type"],
            f"// existing {inc['type']} schema is missing required properties: "
            f"{', '.join(inc.get('missing_properties', []))}",
        )
        template_blocks.append(tpl)
    templates_str = DIVIDER.join(template_blocks)

    return (url, present, missing_str, templates_str)


def _style_workbook(ws, n_rows: int) -> None:
    headers = ["Page URL", "Schemas Present", "Schemas Missing / Incomplete", "JSON-LD Template (paste-ready)"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row:
            cell.alignment = wrap

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 80
    ws.row_dimensions[1].height = 28
    for r_idx in range(2, n_rows + 2):
        ws.row_dimensions[r_idx].height = 240


def write_report(results: list[dict], out_path: Path | None = None, append: bool = False) -> Path:
    """
    Build the Excel workbook from a list of audit results.

    append=True opens an existing workbook (if present at out_path) and appends
    rows instead of overwriting. Use this for batch runs across multiple sessions.
    """
    if out_path is None:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"schema-audit-{date.today().isoformat()}.xlsx"

    if append and out_path.exists():
        wb = load_workbook(out_path)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Schema Audit"
        start_row = 2  # header is row 1

    for i, r in enumerate(results):
        row_data = _row_for(r)
        for col, val in enumerate(row_data, start=1):
            ws.cell(row=start_row + i, column=col, value=val)

    total_rows = ws.max_row - 1
    _style_workbook(ws, total_rows)
    wb.save(out_path)
    return out_path


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: python report.py <results.json> [--append] [--out path.xlsx]", file=sys.stderr)
        return 2
    results_path = Path(sys.argv[1])
    append = "--append" in sys.argv
    out_path: Path | None = None
    if "--out" in sys.argv:
        out_path = Path(sys.argv[sys.argv.index("--out") + 1])

    results = json.loads(results_path.read_text(encoding="utf-8"))
    written = write_report(results, out_path=out_path, append=append)
    print(f"Wrote {written}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
